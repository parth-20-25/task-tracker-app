import unittest
import base64
from io import BytesIO
from openpyxl import Workbook

try:
    from openpyxl.drawing.image import Image as OpenpyxlImage
except Exception:  # pragma: no cover - local env may omit Pillow
    OpenpyxlImage = None

try:
    from fastapi.testclient import TestClient
except Exception:  # pragma: no cover - local env may omit httpx
    TestClient = None

from app.main import app, build_rows, detect_header_hints, find_metadata_row, _process_workbook


PNG_BYTES = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII="
)


def add_test_image(worksheet, cell: str):
    if OpenpyxlImage is None:
        raise unittest.SkipTest("openpyxl image tests require Pillow")
    worksheet.add_image(OpenpyxlImage(BytesIO(PNG_BYTES)), cell)


class SemanticFixtureParserTests(unittest.TestCase):
    def test_build_rows_parses_shifted_semantic_fixture_row(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet["A1"] = "WBS-PARC2600M001-Fuel Tank weld Line_CLIENT_ONE"
        worksheet["B3"] = "Part Name"
        worksheet["C3"] = "QTY"
        worksheet["D3"] = "Image"
        worksheet["E3"] = "Fixture Type"
        worksheet["G3"] = "OP.NO"
        worksheet["H3"] = "Fixture No"
        worksheet["B4"] = "STIFFNER MTG BKT LH/RH SUB ASSLY"
        worksheet["C4"] = 2
        worksheet["D4"] = ""
        worksheet["E4"] = "Robotic MIG Welding fixture"
        worksheet["G4"] = "OP 11"
        worksheet["H4"] = "PARC26001001"

        metadata_row, _ = find_metadata_row(worksheet)
        header_hints = detect_header_hints(worksheet, metadata_row)
        rows, errors = build_rows(worksheet, metadata_row, header_hints, {})

        self.assertEqual(errors, [])
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["fixture_no"], "PARC26001001")
        self.assertEqual(rows[0]["op_no"], "OP 11")
        self.assertEqual(rows[0]["part_name"], "STIFFNER MTG BKT LH/RH SUB ASSLY")
        self.assertEqual(rows[0]["fixture_type"], "Robotic MIG Welding fixture")
        self.assertEqual(rows[0]["qty"], "2")
        self.assertEqual(rows[0]["parser_confidence"], "HIGH")

    def test_build_rows_preserves_letter_suffix_in_op_number(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet["A1"] = "WBS-PARC2600M001-Fuel Tank weld Line_CLIENT_ONE"
        worksheet["A3"] = "Fixture No"
        worksheet["B3"] = "OP.NO"
        worksheet["C3"] = "Part Name"
        worksheet["D3"] = "Fixture Type"
        worksheet["E3"] = "QTY"
        worksheet["A4"] = "PARC26001009"
        worksheet["B4"] = "Op. 10AB"
        worksheet["C4"] = "SUFFIX PART"
        worksheet["D4"] = "Checking fixture"
        worksheet["E4"] = 1

        metadata_row, _ = find_metadata_row(worksheet)
        header_hints = detect_header_hints(worksheet, metadata_row)
        rows, errors = build_rows(worksheet, metadata_row, header_hints, {})

        self.assertEqual(errors, [])
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["op_no"], "OP 10AB")

    def test_build_rows_skips_title_and_header_rows(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet["A1"] = "WBS-PARC2600M001-Fuel Tank weld Line_CLIENT_ONE"
        worksheet["A2"] = "Fuel Tank Weld Line"
        worksheet["A3"] = "FIXTURE NO"
        worksheet["B3"] = "OP.NO"
        worksheet["C3"] = "Part Name"
        worksheet["D3"] = "Fixture Type"
        worksheet["E3"] = "QTY"
        worksheet["A4"] = "PARC26001002"
        worksheet["B4"] = "OP 12"
        worksheet["C4"] = "INNER BRACKET SUB ASSLY"
        worksheet["D4"] = "Checking fixture"
        worksheet["E4"] = 1

        metadata_row, _ = find_metadata_row(worksheet)
        header_hints = detect_header_hints(worksheet, metadata_row)
        rows, errors = build_rows(worksheet, metadata_row, header_hints, {})

        self.assertEqual(errors, [])
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["fixture_no"], "PARC26001002")

    def test_build_rows_rejects_low_information_candidate_rows(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet["A1"] = "WBS-PARC2600M001-Fuel Tank weld Line_CLIENT_ONE"
        worksheet["A3"] = "PARC26001003"
        worksheet["B3"] = "OP 13"

        metadata_row, _ = find_metadata_row(worksheet)
        header_hints = detect_header_hints(worksheet, metadata_row)
        rows, errors = build_rows(worksheet, metadata_row, header_hints, {})

        self.assertEqual(rows, [])
        self.assertEqual(len(errors), 1)
        self.assertIn("Missing required Excel headers", errors[0]["error_message"])

    def test_build_rows_skips_non_fixture_garbage_rows(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet["A1"] = "WBS-PARC2600M001-Fuel Tank weld Line_CLIENT_ONE"
        worksheet["A2"] = "Merged title style row"
        worksheet["B3"] = "Another decorative subtitle"

        metadata_row, _ = find_metadata_row(worksheet)
        header_hints = detect_header_hints(worksheet, metadata_row)
        rows, errors = build_rows(worksheet, metadata_row, header_hints, {})

        self.assertEqual(rows, [])
        self.assertEqual(len(errors), 1)
        self.assertIn("Missing required Excel headers", errors[0]["error_message"])

    def test_build_rows_rejects_ambiguous_part_name_without_guessing(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet["A1"] = "WBS-PARC2600M001-Fuel Tank weld Line_CLIENT_ONE"
        worksheet["A3"] = "PARC26001004"
        worksheet["B3"] = "OP 14"
        worksheet["C3"] = "LH BRACKET"
        worksheet["D3"] = "RH BRACKET"
        worksheet["E3"] = "Checking fixture"
        worksheet["F3"] = 2
        worksheet["G3"] = "PARC scope"

        metadata_row, _ = find_metadata_row(worksheet)
        header_hints = detect_header_hints(worksheet, metadata_row)
        rows, errors = build_rows(worksheet, metadata_row, header_hints, {})

        self.assertEqual(rows, [])
        self.assertEqual(len(errors), 1)
        self.assertIn("Missing required Excel headers", errors[0]["error_message"])

    def test_build_rows_carries_vertical_merged_fixture_values(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet["A1"] = "WBS-PARC2600M001-Fuel Tank weld Line_CLIENT_ONE"
        worksheet["A3"] = "Fixture No"
        worksheet["B3"] = "OP.NO"
        worksheet["C3"] = "Part Name"
        worksheet["D3"] = "Fixture Type"
        worksheet["E3"] = "QTY"
        worksheet["A4"] = "PARC26001005"
        worksheet.merge_cells("A4:A5")
        worksheet["B4"] = "OP 21"
        worksheet["C4"] = "FIRST PART"
        worksheet["D4"] = "Checking fixture"
        worksheet["E4"] = 1
        worksheet["B5"] = "OP 22"
        worksheet["C5"] = "SECOND PART"
        worksheet["D5"] = "Checking fixture"
        worksheet["E5"] = 2

        metadata_row, _ = find_metadata_row(worksheet)
        header_hints = detect_header_hints(worksheet, metadata_row)
        rows, errors = build_rows(worksheet, metadata_row, header_hints, {})

        self.assertEqual(errors, [])
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1]["fixture_no"], "PARC26001005")
        self.assertEqual(rows[1]["op_no"], "OP 22")
        self.assertEqual(rows[1]["qty"], "2")

    def test_build_rows_ignores_remark_column_for_validation(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet["A1"] = "WBS-PARC2600M001-Fuel Tank weld Line_CLIENT_ONE"
        worksheet["A3"] = "Fixture No"
        worksheet["B3"] = "OP.NO"
        worksheet["C3"] = "Part Name"
        worksheet["D3"] = "Fixture Type"
        worksheet["E3"] = "QTY"
        worksheet["F3"] = "Remark"
        worksheet["A4"] = "PARC26001006"
        worksheet["B4"] = "OP 31"
        worksheet["C4"] = "REMARK SAFE PART"
        worksheet["D4"] = "Custom type text"
        worksheet["E4"] = 1
        worksheet["F4"] = ""
        worksheet["A5"] = "PARC26001007"
        worksheet["B5"] = "OP 32"
        worksheet["C5"] = "REMARK FREE PART"
        worksheet["D5"] = "Another custom type"
        worksheet["E5"] = 2
        worksheet["F5"] = "Any invalid free text here"

        metadata_row, _ = find_metadata_row(worksheet)
        header_hints = detect_header_hints(worksheet, metadata_row)
        rows, errors = build_rows(worksheet, metadata_row, header_hints, {})

        self.assertEqual(errors, [])
        self.assertEqual(len(rows), 2)
        self.assertIsNone(rows[0]["remark"])
        self.assertIsNone(rows[1]["remark"])

    def test_detect_header_hints_rejects_loose_renamed_columns(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet["A1"] = "WBS-PARC2600M001-Fuel Tank weld Line_CLIENT_ONE"
        worksheet["A3"] = "Fixture Identifier"
        worksheet["B3"] = "Operation"
        worksheet["C3"] = "Description"
        worksheet["D3"] = "Type"
        worksheet["E3"] = "Nos"

        metadata_row, _ = find_metadata_row(worksheet)
        header_hints = detect_header_hints(worksheet, metadata_row)
        rows, errors = build_rows(worksheet, metadata_row, header_hints, {})

        self.assertEqual(rows, [])
        self.assertEqual(len(errors), 1)
        self.assertIn("Missing required Excel headers", errors[0]["error_message"])

    def test_process_workbook_uses_fixture_sheet_not_first_cost_sheet(self):
        workbook = Workbook()
        cost_sheet = workbook.active
        cost_sheet.title = "SA 30"
        cost_sheet["A1"] = "Robotic MIG Welding Fixture"
        cost_sheet["A2"] = "Sr. No."
        cost_sheet["D2"] = "Qty"
        cost_sheet["A3"] = 1
        cost_sheet["D3"] = 8

        fixture_sheet = workbook.create_sheet("WBS-PARC2600M001")
        fixture_sheet["A1"] = "WBS-PARC2600M001-Fuel Tank weld Line_CLIENT_ONE"
        fixture_sheet["A3"] = "Fixture No"
        fixture_sheet["B3"] = "OP.NO"
        fixture_sheet["C3"] = "Part Name"
        fixture_sheet["D3"] = "Fixture Type"
        fixture_sheet["E3"] = "QTY"
        fixture_sheet["A4"] = "PARC26001008"
        fixture_sheet["B4"] = "10.0"
        fixture_sheet["C4"] = "SPM STATION"
        fixture_sheet["D4"] = "Checking fixture"
        fixture_sheet["E4"] = 2

        buffer = BytesIO()
        workbook.save(buffer)

        result = _process_workbook(buffer.getvalue())

        self.assertEqual(result["file_info"]["project_code"], "PARC2600M001")
        self.assertEqual(len(result["rows"]), 1)
        self.assertEqual(result["rows"][0]["fixture_no"], "PARC26001008")
        self.assertEqual(result["rows"][0]["op_no"], "OP 10")
        self.assertEqual(result["rows"][0]["qty"], "2")
        self.assertEqual(result["errors"], [])

    def test_process_workbook_rejects_invalid_op_number_with_letters_before_digits(self):
        workbook = Workbook()
        fixture_sheet = workbook.active
        fixture_sheet.title = "WBS-PARC2600M001"
        fixture_sheet["A1"] = "WBS-PARC2600M001-Fuel Tank weld Line_CLIENT_ONE"
        fixture_sheet["A3"] = "Fixture No"
        fixture_sheet["B3"] = "OP.NO"
        fixture_sheet["C3"] = "Part Name"
        fixture_sheet["D3"] = "Fixture Type"
        fixture_sheet["E3"] = "QTY"
        fixture_sheet["A4"] = "PARC26001010"
        fixture_sheet["B4"] = "OP A10"
        fixture_sheet["C4"] = "INVALID OP PART"
        fixture_sheet["D4"] = "Checking fixture"
        fixture_sheet["E4"] = 1

        buffer = BytesIO()
        workbook.save(buffer)

        result = _process_workbook(buffer.getvalue())

        self.assertEqual(result["rows"], [])
        self.assertEqual(len(result["errors"]), 1)
        self.assertIn("Invalid OP.NO format", result["errors"][0]["error_message"])

    def test_process_workbook_parses_wbs_with_dash_only_company_separator(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet["A1"] = "WBS - PARC2600M001 - Fuel Tank weld Line - Belrise Industries Limited"
        worksheet["A3"] = "Sr No"
        worksheet["B3"] = "Fixture Number"
        worksheet["C3"] = "Operation No"
        worksheet["D3"] = "Part Name"
        worksheet["E3"] = "Image"
        worksheet["F3"] = "Main Image"
        worksheet["G3"] = "Fixture Type"
        worksheet["H3"] = "QTY"
        worksheet["I3"] = "Remarks"
        worksheet["A4"] = 1
        worksheet["B4"] = "PARC26001011"
        worksheet["C4"] = "OP 11"
        worksheet["D4"] = "FUEL TANK"
        worksheet["E4"] = "#VALUE!"
        worksheet["F4"] = "#VALUE!"
        worksheet["G4"] = "Checking fixture"
        worksheet["H4"] = 1
        worksheet["I4"] = "Customer Scope"

        buffer = BytesIO()
        workbook.save(buffer)

        result = _process_workbook(buffer.getvalue())

        self.assertEqual(result["file_info"]["project_code"], "PARC2600M001")
        self.assertEqual(result["file_info"]["project_name"], "Fuel Tank weld Line")
        self.assertEqual(result["file_info"]["company_name"], "Belrise Industries Limited")
        self.assertEqual(len(result["rows"]), 1)
        self.assertIsNone(result["rows"][0]["image_1_url"])
        self.assertIsNone(result["rows"][0]["image_1_upload"])

    def test_process_workbook_format_a_uses_rightmost_main_image_column(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet["A1"] = "WBS-PARC2600M001_Fuel Tank weld Line_Belrise Industries Limited_Pune"
        worksheet["B3"] = "SR NO"
        worksheet["C3"] = "FIXTURE NO"
        worksheet["D3"] = "OP NO."
        worksheet["E3"] = "PART NAME"
        worksheet["F3"] = "Legacy Image"
        worksheet["G3"] = "Main Image"
        worksheet["H3"] = "FIXTURE TYPE"
        worksheet["I3"] = "QTY"
        worksheet["J3"] = "Remarks"
        worksheet["B4"] = 1
        worksheet["C4"] = "PARC26001012"
        worksheet["D4"] = "OP 12"
        worksheet["E4"] = "FORMAT A PART"
        worksheet["F4"] = "#VALUE!"
        worksheet["H4"] = "Checking fixture"
        worksheet["I4"] = 1
        worksheet["J4"] = "PARC Scope"
        add_test_image(worksheet, "G4")

        buffer = BytesIO()
        workbook.save(buffer)

        result = _process_workbook(buffer.getvalue())

        self.assertEqual(result["file_info"]["company_name"], "Belrise Industries Limited")
        self.assertEqual(len(result["rows"]), 1)
        row = result["rows"][0]
        self.assertIsNotNone(row["image_1_upload"])
        self.assertEqual(row["image_1_upload"]["anchor"]["column"], 7)
        self.assertIsNone(row["remark"])

    def test_process_workbook_format_b_uses_rightmost_main_image_column(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet["A1"] = "WBS-PARC2600M001 - Fuel Tank weld Line_ Belrise Industries Limited_ Pune"
        worksheet["A3"] = "Sr No"
        worksheet["B3"] = "Fixture No"
        worksheet["C3"] = "Op No"
        worksheet["D3"] = "Part Name"
        worksheet["E3"] = "Legacy Image"
        worksheet["F3"] = "Main Image"
        worksheet["G3"] = "Fixture Type"
        worksheet["H3"] = "Qty"
        worksheet["I3"] = "Remarks"
        worksheet["A4"] = 1
        worksheet["B4"] = "PARC26001013"
        worksheet["C4"] = "OP 13"
        worksheet["D4"] = "FORMAT B PART"
        worksheet["E4"] = "#VALUE!"
        worksheet["G4"] = "Checking fixture"
        worksheet["H4"] = 1
        worksheet["I4"] = "Free text that must not affect parsing"
        add_test_image(worksheet, "F4")

        buffer = BytesIO()
        workbook.save(buffer)

        result = _process_workbook(buffer.getvalue())

        self.assertEqual(len(result["rows"]), 1)
        row = result["rows"][0]
        self.assertEqual(row["fixture_no"], "PARC26001013")
        self.assertIsNotNone(row["image_1_upload"])
        self.assertEqual(row["image_1_upload"]["anchor"]["column"], 6)
        self.assertNotIn("#VALUE!", row["image_1_upload"]["content_base64"])


class ExtractEndpointValidationTests(unittest.TestCase):
    def setUp(self):
        if TestClient is None:
            self.skipTest("fastapi testclient dependencies are not installed")
        import app.main as main_module

        self.original_token = main_module.SERVICE_TOKEN
        main_module.SERVICE_TOKEN = "test-token"
        self.client = TestClient(app)

    def tearDown(self):
        import app.main as main_module

        main_module.SERVICE_TOKEN = self.original_token

    def test_extract_rejects_wrong_form_field_name_with_422(self):
        response = self.client.post(
            "/extract",
            headers={"x-extraction-token": "test-token"},
            files={
                "excel": (
                    "fixtures.xlsx",
                    b"not-a-real-workbook",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        self.assertEqual(response.status_code, 422)
        self.assertIn("Expected multipart field 'file'", response.json()["errors"][0]["error_message"])


if __name__ == "__main__":
    unittest.main()
