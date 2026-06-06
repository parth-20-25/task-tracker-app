from __future__ import annotations

import asyncio
import base64
import json
import logging
import os
import re
import time
from io import BytesIO
from pathlib import Path
from typing import Any

from fastapi import FastAPI, Header, Request
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from starlette.datastructures import UploadFile as StarletteUploadFile


def normalize_runtime_mode(value: str | None) -> str:
    return "production" if (value or "").strip().lower() == "production" else "development"


def unquote_env_value(value: str) -> str:
    stripped_value = value.strip()
    if len(stripped_value) >= 2 and stripped_value[0] == stripped_value[-1] and stripped_value[0] in {"'", '"'}:
        return stripped_value[1:-1]
    return stripped_value


def load_service_env() -> None:
    mode = normalize_runtime_mode(os.getenv("PYTHON_ENV") or os.getenv("ENV") or os.getenv("NODE_ENV"))
    os.environ["NODE_ENV"] = mode

    service_root = Path(__file__).resolve().parents[1]
    candidate_files = [service_root / f".env.{mode}"]
    if mode == "development":
        candidate_files.append(service_root / ".env")

    for env_path in candidate_files:
        if not env_path.exists():
            continue

        for raw_line in env_path.read_text(encoding="utf-8").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue

            key, value = line.split("=", 1)
            key = key.strip()
            if key:
                os.environ.setdefault(key, unquote_env_value(value))
        break


load_service_env()

ALLOWED_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
ALLOWED_EXTENSION = ".xlsx"
EXPECTED_UPLOAD_FIELD = "file"
MAX_UPLOAD_BYTES = int(os.getenv("EXTRACTION_MAX_UPLOAD_BYTES", str(10 * 1024 * 1024)))
SERVICE_TOKEN = (os.getenv("EXTRACTION_SERVICE_TOKEN") or os.getenv("DESIGN_EXTRACTION_SERVICE_TOKEN") or "").strip()
IMAGE_SLOT_NAME = "image_1_url"

FIXTURE_NUMBER_PATTERN = re.compile(r"^PARC\d{4,}$", re.IGNORECASE)
OP_NUMBER_PATTERN = re.compile(r"^OP\.?\s*\d+[A-Z]*$", re.IGNORECASE)
FIXTURE_TYPE_KEYWORDS = (
    "fixture",
    "weld",
    "welding",
    "mig",
    "tig",
    "robotic",
    "robot",
    "assy",
    "assly",
    "jig",
    "gauge",
    "check",
    "checking",
    "inspection",
    "holding",
    "clamping",
    "mounting",
)
HEADER_FIELD_ALIASES = {
    "row_reference": {
        "sno",
        "srno",
        "serialno",
        "serialnumber",
    },
    "fixture_no": {
        "fixtureno",
        "fixturenumber",
    },
    "op_no": {
        "opno",
        "operationno",
    },
    "part_name": {
        "partname",
    },
    "fixture_type": {
        "fixturetype",
    },
    "qty": {"qty", "quantity"},
    "remark": {"remark", "remarks"},
}
REQUIRED_HEADER_FIELDS = ("fixture_no", "op_no", "part_name", "fixture_type", "qty")
EXPECTED_HEADER_LABELS = {
    "fixture_no": "Fixture No",
    "op_no": "OP.NO",
    "part_name": "Part Name",
    "fixture_type": "Fixture Type",
    "qty": "QTY",
}
METADATA_FIELD_ALIASES = {
    "project_code": {"wbsprojectno", "wbsprojectnumber", "projectno", "projectnumber"},
    "project_name": {"projectname"},
    "company_name": {"companyname", "customername"},
}

logger = logging.getLogger("design_extraction")
if not logger.handlers:
    logging.basicConfig(level=os.getenv("LOG_LEVEL", "DEBUG").upper(), format="%(asctime)s - %(message)s")

# Performance timing context manager
class Timer:
    def __init__(self, name: str):
        self.name = name
        self.start = None

    def __enter__(self):
        self.start = time.perf_counter()
        return self

    def __exit__(self, *args):
        elapsed = time.perf_counter() - self.start
        log_event("timing", operation=self.name, elapsed_ms=round(elapsed * 1000, 2))

app = FastAPI()


@app.api_route("/", methods=["GET", "HEAD"])
def root() -> dict[str, str]:
    return {"status": "ok"}


@app.api_route("/health", methods=["GET", "HEAD"])
def health() -> dict[str, str]:
    return {"status": "ok"}


def log_event(event: str, **payload: Any) -> None:
    record = {"event": event, **payload}
    logger.info(json.dumps(record, default=str))


def get_database_connection():
    database_url = os.getenv("DATABASE_URL")
    if not database_url:
        return None

    import psycopg2

    return psycopg2.connect(database_url, sslmode="require")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_wbs_header_text(value: Any) -> str:
    """
    Normalize WBS header text for robust matching.
    Handles variations like:
    - WBS - PARC2600M001-Fuel Tank weld Line
    - WBS- PARC2600M001 Fuel Tank weld Line
    - WBS - PARC2600M001_Fuel Tank weld Line
    - WBS_PARC2600M001 - Fuel Tank weld Line
    - WBS -PARC2600M001- Fuel Tank weld Line
    - WBS - PARC2600M001-Fuel Tank weld Line_Belrise Industries Limited
    """
    if value is None:
        return ""

    text = str(value).strip()

    text = re.sub(r"\s+", " ", text)
    return re.sub(r"^WBS\s*[-_]?\s*", "WBS-", text, flags=re.IGNORECASE)


def normalize_header(value: Any) -> str:
    return "".join(ch for ch in normalize_text(value).lower() if ch.isalnum())


def normalize_key(value: Any) -> str:
    return " ".join(normalize_text(value).lower().split())


def normalize_fixture_number(value: Any) -> str:
    return normalize_text(value).upper().replace(" ", "")


def build_error(message: str, excel_row: int | None = None, raw_data: dict[str, Any] | None = None) -> dict[str, Any]:
    return {
        "excel_row": excel_row,
        "error_message": message,
        "raw_data": raw_data or {},
    }


def build_error_response(status_code: int, message: str, errors: list[dict[str, Any]] | None = None) -> JSONResponse:
    return JSONResponse(
        status_code=status_code,
        content={
            "message": message,
            "errors": errors or [],
        },
    )


def parse_wbs_header(raw_header: str) -> dict[str, str]:
    """
    Parse WBS header with flexible separator support.
    Handles formats like:
    - WBS-PARC2500M119_Oil Retainer Parts_Belrise Industries LTD_Pune
    - WBS-PARC2500M119 - Oil Retainer Parts_ Belrise Industries LTD_ Pune
    - WBS - PARC2600M001-Fuel Tank weld Line_Belrise Industries Limited
    - WBS - PARC2600M001 - Fuel Tank weld Line - Belrise Industries Limited
    """
    header = re.sub(r"\s+", " ", str(raw_header).strip())

    if not header.upper().startswith("WBS"):
        raise ValueError("Invalid header format: missing 'WBS' prefix.")

    wbs_code_match = re.search(r"\bWBS\b\s*[-_]?\s*(PARC[A-Z0-9]+)(?=\s|[-_]|$)", header, re.IGNORECASE)
    if not wbs_code_match:
        raise ValueError("Invalid header format: could not extract project code.")

    project_code = wbs_code_match.group(1).upper()
    remainder = header[wbs_code_match.end(1):].strip()
    remainder = re.sub(r"^[-_\s]+", "", remainder)

    if not remainder:
        raise ValueError("Invalid WBS format: expected project name and company name after Project No.")

    if "_" in remainder:
        parts = [part.strip(" -_") for part in re.split(r"\s*_\s*", remainder) if part.strip(" -_")]
    else:
        parts = [part.strip(" -_") for part in re.split(r"\s+-\s+", remainder) if part.strip(" -_")]

    if len(parts) < 2:
        raise ValueError("Invalid WBS format: expected WBS-<Project No>-<Project Name>_<Company Name>.")

    project_name = re.sub(r"\s+", " ", re.sub(r"[_]+", " ", parts[0])).strip()
    company_name = re.sub(r"\s+", " ", re.sub(r"[_]+", " ", parts[1])).strip()

    if not project_name:
        raise ValueError("Invalid WBS format: Project Name is required.")

    if not company_name:
        raise ValueError("Invalid WBS format: Company Name is required.")

    return {
        "project_code": project_code,
        "project_name": project_name,
        "company_name": company_name,
    }


def find_metadata_row(worksheet) -> tuple[int, str]:
    """
    Find the WBS metadata row using robust pattern matching.
    Supports flexible formatting: WBS -, WBS-, WBS_, with various spacing.
    """
    # WBS pattern: PARC followed by 4+ digits and optional M + digits
    wbs_code_pattern = re.compile(r"PARC\d{4,}[A-Z0-9]*", re.IGNORECASE)

    for row_index, row_values in enumerate(worksheet.iter_rows(min_row=1, max_row=50, values_only=True), start=1):
        for cell_value in row_values:
            if cell_value is None:
                continue

            # Normalize the cell text for flexible matching
            normalized = normalize_wbs_header_text(cell_value)

            # Check if normalized text starts with WBS- and contains a PARC code
            if normalized.startswith("WBS-") and wbs_code_pattern.search(normalized):
                return row_index, normalized

            # Fallback: Check raw text for WBS-like patterns
            raw_text = normalize_text(cell_value).upper()
            if raw_text.startswith(("WBS", "WBS-", "WBS_", "WBS ")) and wbs_code_pattern.search(raw_text):
                # Normalize and return
                return row_index, normalize_wbs_header_text(cell_value)

    labeled_values: dict[str, str] = {}
    labeled_rows: list[int] = []
    for row_index, row_values in enumerate(worksheet.iter_rows(min_row=1, max_row=50, values_only=True), start=1):
        normalized_values = [normalize_text(value) for value in row_values]
        for column_index, cell_text in enumerate(normalized_values):
            if not cell_text:
                continue

            label_text = cell_text
            inline_value = ""
            if ":" in cell_text:
                label_text, inline_value = [part.strip() for part in cell_text.split(":", 1)]

            normalized_label = normalize_header(label_text)
            field_name = next(
                (
                    candidate_field
                    for candidate_field, aliases in METADATA_FIELD_ALIASES.items()
                    if normalized_label in aliases
                ),
                None,
            )
            if not field_name or field_name in labeled_values:
                continue

            next_value = inline_value
            if not next_value:
                for candidate_value in normalized_values[column_index + 1:]:
                    if candidate_value:
                        next_value = candidate_value
                        break

            if next_value:
                labeled_values[field_name] = next_value
                labeled_rows.append(row_index)

    if all(field_name in labeled_values for field_name in ("project_code", "project_name", "company_name")):
        project_code = normalize_fixture_number(labeled_values["project_code"])
        if not wbs_code_pattern.fullmatch(project_code):
            raise ValueError("Invalid WBS format: Project No must be a PARC project code.")

        metadata_value = f"WBS-{project_code}-{labeled_values['project_name']}_{labeled_values['company_name']}"
        return min(labeled_rows) if labeled_rows else 1, metadata_value

    raise ValueError("Could not find the WBS metadata row in the workbook.")


def find_workbook_metadata(workbook) -> tuple[str, int, str]:
    for worksheet in workbook.worksheets:
        try:
            metadata_row, metadata_value = find_metadata_row(worksheet)
            return worksheet.title, metadata_row, metadata_value
        except ValueError:
            continue

    raise ValueError("Could not find the WBS metadata row in the workbook.")


IMAGE_MIME_TYPES = {
    "bmp": "image/bmp",
    "gif": "image/gif",
    "heic": "image/heic",
    "heif": "image/heif",
    "jpeg": "image/jpeg",
    "jpg": "image/jpeg",
    "png": "image/png",
    "webp": "image/webp",
}


def normalize_image_extension(value: Any) -> str:
    extension = normalize_text(value).lower().replace(".", "")
    if extension == "jpeg":
        return "jpg"
    if extension in IMAGE_MIME_TYPES:
        return extension
    return "png"


def build_image_payload(image_bytes: bytes, excel_row: int, excel_column: int, slot_name: str, extension: str) -> dict[str, Any]:
    normalized_extension = normalize_image_extension(extension)
    return {
        "content_base64": base64.b64encode(image_bytes).decode("ascii"),
        "mime_type": IMAGE_MIME_TYPES.get(normalized_extension, "image/png"),
        "extension": normalized_extension,
        "anchor": {
            "row": excel_row,
            "column": excel_column,
            "slot": slot_name,
        },
    }


def get_main_image_column(header_hints: dict[str, Any]) -> int | None:
    value = header_hints.get("__main_image_column")
    try:
        column = int(value)
    except (TypeError, ValueError):
        return None
    return column if column > 0 else None


def get_image_header_columns(header_hints: dict[str, Any]) -> set[int]:
    columns = header_hints.get("__image_columns")
    if not isinstance(columns, list):
        return set()
    resolved: set[int] = set()
    for column in columns:
        try:
            column_number = int(column)
        except (TypeError, ValueError):
            continue
        if column_number > 0:
            resolved.add(column_number)
    return resolved


def resolve_image_slot_for_anchor(excel_column: int, header_hints: dict[str, Any]) -> str | None:
    main_image_column = get_main_image_column(header_hints)
    if not main_image_column:
        return None

    if excel_column == main_image_column:
        return IMAGE_SLOT_NAME

    image_header_columns = get_image_header_columns(header_hints)
    if abs(excel_column - main_image_column) <= 1 and excel_column not in image_header_columns:
        return IMAGE_SLOT_NAME

    return None


def extract_anchored_images(worksheet, header_hints: dict[str, Any] | None = None) -> tuple[dict[int, dict[str, dict[str, Any]]], list[dict[str, Any]]]:
    header_hints = header_hints or {}
    images_by_row: dict[int, dict[str, dict[str, Any]]] = {}
    errors: list[dict[str, Any]] = []
    main_image_column = get_main_image_column(header_hints)
    image_header_columns = sorted(get_image_header_columns(header_hints))

    # DEBUG: Log worksheet image detection details
    log_event("image_extraction_start", sheet_title=worksheet.title)

    # Comprehensive image source detection for different Excel versions
    all_images = []
    image_sources = []

    # Source 1: Direct _images attribute (openpyxl 3.1.5+)
    direct_images = getattr(worksheet, "_images", [])
    if direct_images:
        all_images.extend(direct_images)
        image_sources.append(f"_images({len(direct_images)})")
        log_event("image_source_direct", count=len(direct_images))
        # Debug first image structure
        if direct_images:
            first_img = direct_images[0]
            log_event("first_image_debug",
                      image_type=type(first_img).__name__,
                      has_anchor=hasattr(first_img, "anchor"),
                      anchor_type=type(getattr(first_img, "anchor", None)).__name__,
                      image_attrs=[a for a in dir(first_img) if not a.startswith('_')][:15])

    # Source 2: _drawing._images (alternative location)
    drawings = getattr(worksheet, "_drawing", None)
    if drawings:
        drawing_images = getattr(drawings, "_images", [])
        if drawing_images:
            all_images.extend(drawing_images)
            image_sources.append(f"_drawing._images({len(drawing_images)})")
            log_event("image_source_drawing", count=len(drawing_images))

    # Source 3: _charts (sometimes images are stored as charts)
    if drawings:
        drawing_charts = getattr(drawings, "_charts", [])
        chart_images = [chart for chart in drawing_charts if hasattr(chart, 'anchor') and hasattr(chart, '_data')]
        if chart_images:
            all_images.extend(chart_images)
            image_sources.append(f"_drawing._charts_images({len(chart_images)})")
            log_event("image_source_charts", count=len(chart_images))

    # Source 4: worksheet._drawings (openpyxl alternative)
    worksheet_drawings = getattr(worksheet, "_drawings", [])
    if worksheet_drawings:
        drawing_images_from_ws = []
        for drawing in worksheet_drawings:
            if hasattr(drawing, "_images"):
                drawing_images_from_ws.extend(drawing._images)
        if drawing_images_from_ws:
            all_images.extend(drawing_images_from_ws)
            image_sources.append(f"_drawings._images({len(drawing_images_from_ws)})")
            log_event("image_source_worksheet_drawings", count=len(drawing_images_from_ws))

    # Log comprehensive detection results
    log_event("image_detection_comprehensive",
              sheet_title=worksheet.title,
              total_images_found=len(all_images),
              sources=image_sources,
              main_image_column=main_image_column,
              image_header_columns=image_header_columns,
              worksheet_attrs=[attr for attr in dir(worksheet) if not attr.startswith('__') and any(keyword in attr.lower() for keyword in ['image', 'drawing', 'chart'])][:10])

    if not all_images:
        log_event("no_images_found", sheet_title=worksheet.title)
        return images_by_row, errors

    for idx, image in enumerate(all_images):
        anchor = getattr(image, "anchor", None)

        # Handle different anchor types
        anchor_from = None
        anchor_type = type(anchor).__name__ if anchor else None

        if anchor is None:
            log_event("image_no_anchor", image_index=idx, image_type=type(image).__name__)
            errors.append(build_error("Found an image without an anchor.", excel_row=None))
            continue

        # Get _from attribute from anchor
        anchor_from = getattr(anchor, "_from", None)

        # Handle different anchor structures
        if anchor_from is None:
            # Try direct row/col on anchor (for some Excel versions)
            if hasattr(anchor, "row") and hasattr(anchor, "col"):
                anchor_from = anchor
            else:
                log_event("image_anchor_missing",
                          image_index=idx,
                          image_type=type(image).__name__,
                          anchor_type=anchor_type,
                          anchor_attrs=[a for a in dir(anchor) if not a.startswith('_')][:10])
                errors.append(build_error(
                    f"Found an image without a readable anchor position. Anchor type: {anchor_type}",
                    excel_row=None
                ))
                continue

        # Extract row and column from anchor with comprehensive fallback handling
        try:
            row_val = None
            col_val = None
            
            # Method 1: Standard AnchorMarker with _from attribute
            if hasattr(anchor_from, "row") and hasattr(anchor_from, "col"):
                row_val = anchor_from.row
                col_val = anchor_from.col
                log_event("anchor_method_1_success", image_index=idx, row=row_val, col=col_val)
            else:
                # Method 2: Direct anchor attributes (some Excel versions)
                if hasattr(anchor, "row") and hasattr(anchor, "col"):
                    row_val = anchor.row
                    col_val = anchor.col
                    log_event("anchor_method_2_success", image_index=idx, row=row_val, col=col_val)
                else:
                    # Method 3: Try to get from anchor._from._from (nested structure)
                    nested_from = getattr(anchor_from, "_from", None) if anchor_from else None
                    if nested_from and hasattr(nested_from, "row") and hasattr(nested_from, "col"):
                        row_val = nested_from.row
                        col_val = nested_from.col
                        log_event("anchor_method_3_success", image_index=idx, row=row_val, col=col_val)
                    else:
                        # Method 4: Try anchor.from attribute (alternative naming)
                        direct_from = getattr(anchor, "from", None)
                        if direct_from and hasattr(direct_from, "row") and hasattr(direct_from, "col"):
                            row_val = direct_from.row
                            col_val = direct_from.col
                            log_event("anchor_method_4_success", image_index=idx, row=row_val, col=col_val)
                        else:
                            # Method 5: Try to extract from anchor properties using inspection
                            anchor_props = {}
                            for prop_name in ['row', 'col', '_row', '_col', 'rowIndex', 'colIndex']:
                                if hasattr(anchor, prop_name):
                                    try:
                                        anchor_props[prop_name] = getattr(anchor, prop_name)
                                    except:
                                        pass
                            
                            if 'row' in anchor_props and 'col' in anchor_props:
                                row_val = anchor_props['row']
                                col_val = anchor_props['col']
                                log_event("anchor_method_5_success", image_index=idx, row=row_val, col=col_val)
                            elif '_row' in anchor_props and '_col' in anchor_props:
                                row_val = anchor_props['_row']
                                col_val = anchor_props['_col']
                                log_event("anchor_method_5b_success", image_index=idx, row=row_val, col=col_val)
                            else:
                                log_event("anchor_extraction_failed",
                                          image_index=idx,
                                          anchor_type=anchor_type,
                                          available_props=list(anchor_props.keys()),
                                          anchor_dir_attrs=[a for a in dir(anchor) if not a.startswith('__')][:20])
                                raise ValueError(f"Could not extract row/col from anchor type {anchor_type}")

            # Validate and convert to Excel coordinates
            if row_val is None or col_val is None:
                raise ValueError(f"Extracted None values: row={row_val}, col={col_val}")
                
            excel_row = int(row_val) + 1  # Convert 0-indexed to 1-indexed
            excel_column = int(col_val) + 1  # Convert 0-indexed to 1-indexed
            
            # Sanity check the coordinates
            if excel_row < 1 or excel_column < 1 or excel_row > 100000 or excel_column > 1000:
                log_event("anchor_coordinates_suspicious",
                          image_index=idx,
                          excel_row=excel_row,
                          excel_column=excel_column,
                          original_row=row_val,
                          original_col=col_val)
                
        except (ValueError, TypeError, AttributeError) as e:
            log_event("image_anchor_parse_error",
                      image_index=idx,
                      error=str(e),
                      anchor_type=anchor_type,
                      anchor_from_type=type(anchor_from).__name__ if anchor_from else None,
                      anchor_attrs=dir(anchor_from) if anchor_from else [],
                      anchor_attrs_full=dir(anchor))
            errors.append(build_error(
                f"Could not parse image anchor position: {e}",
                excel_row=None
            ))
            continue

        log_event("image_anchor_debug",
                  image_index=idx,
                  anchor_type=anchor_type,
                  anchor_row=row_val,
                  anchor_col=col_val,
                  excel_row=excel_row,
                  excel_column=excel_column)

        slot_name = resolve_image_slot_for_anchor(excel_column, header_hints)
        if not slot_name:
            log_event("image_column_rejected",
                      excel_row=excel_row,
                      excel_column=excel_column,
                      main_image_column=main_image_column,
                      image_header_columns=image_header_columns)
            continue

        row_images = images_by_row.setdefault(excel_row, {})

        if slot_name in row_images:
            errors.append(
                build_error(
                    "Multiple images were found in the same mapped column for one row.",
                    excel_row=excel_row,
                    raw_data={"column": excel_column},
                )
            )
            continue

        image_format = normalize_text(getattr(image, "format", "")) or "png"
        log_event("image_processing",
                  excel_row=excel_row,
                  slot_name=slot_name,
                  format=image_format,
                  has_data_method=hasattr(image, "_data"))
        try:
            image_bytes = image._data()
            row_images[slot_name] = build_image_payload(image_bytes, excel_row, excel_column, slot_name, image_format.lower())
            log_event("image_extracted_success",
                      excel_row=excel_row,
                      slot_name=slot_name,
                      byte_count=len(image_bytes),
                      extension=normalize_image_extension(image_format))
        except Exception as exc:
            errors.append(
                build_error(
                    "Failed to extract an embedded image.",
                    excel_row=excel_row,
                    raw_data={"details": str(exc)},
                )
            )

    log_event("image_extraction_complete",
              sheet_title=worksheet.title,
              images_mapped=len(images_by_row),
              errors_count=len(errors))
    return images_by_row, errors


def open_excel_workbook(file_bytes: bytes, *, read_only: bool) -> Any:
    return load_workbook(
        filename=BytesIO(file_bytes),
        data_only=True,
        keep_links=False,
        read_only=read_only,
    )


def extract_images_from_workbook(
    file_bytes: bytes,
    header_hints: dict[str, Any] | None = None,
) -> tuple[dict[int, dict[str, dict[str, Any]]], list[dict[str, Any]]]:
    workbook = open_excel_workbook(file_bytes, read_only=False)
    try:
        worksheet = workbook.active
        return extract_anchored_images(worksheet, header_hints or {})
    finally:
        workbook.close()


def tokenize_header(value: Any) -> set[str]:
    text = normalize_text(value).lower().replace("_", " ")
    return {token for token in re.split(r"[^a-z0-9]+", text) if token}


def match_header_field(cell_text: str) -> str | None:
    normalized = normalize_header(cell_text)

    for field_name, aliases in HEADER_FIELD_ALIASES.items():
        if normalized in aliases:
            return field_name

    return None


def is_image_header(cell_text: str) -> bool:
    normalized = normalize_header(cell_text)
    tokens = tokenize_header(cell_text)

    if not normalized:
        return False

    if normalized in {
        "image",
        "images",
        "mainimage",
        "partimage",
        "fixtureimage",
        "referenceimage",
        "refimage",
        "picture",
        "photo",
        "photograph",
    }:
        return True

    return bool(tokens & {"image", "images", "picture", "photo", "photograph"})


def missing_required_headers(header_hints: dict[str, int]) -> list[str]:
    return [
        EXPECTED_HEADER_LABELS[field_name]
        for field_name in REQUIRED_HEADER_FIELDS
        if field_name not in header_hints
    ]


def detect_header_hints(worksheet, metadata_row: int) -> dict[str, int]:
    with Timer("detect_header_hints_internal"):
        best_match_count = 0
        best_mapping: dict[str, int] = {}

        start_row = max(1, metadata_row + 1)
        end_row = min(max(start_row + 30, 30), worksheet.max_row)

        # Pre-compile the header detection to avoid repeated function calls
        for row_index, row_values in enumerate(
            worksheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True),
            start=start_row,
        ):
            current_mapping: dict[str, int] = {}
            image_columns: list[int] = []
            for column_index, cell_value in enumerate(row_values, start=1):
                cell_text = normalize_text(cell_value)
                field_name = match_header_field(cell_text)
                if field_name and field_name not in current_mapping:
                    current_mapping[field_name] = column_index
                if is_image_header(cell_text):
                    image_columns.append(column_index)

            match_count = len([field_name for field_name in current_mapping if not field_name.startswith("__")])
            has_required_headers = all(field_name in current_mapping for field_name in REQUIRED_HEADER_FIELDS)

            has_better_mapping = has_required_headers and (
                match_count > best_match_count
                or (match_count == best_match_count and image_columns and not best_mapping.get("__main_image_column"))
            )

            if has_better_mapping:
                best_match_count = match_count
                current_mapping["__row"] = row_index
                if image_columns:
                    current_mapping["__image_columns"] = image_columns
                    current_mapping["__main_image_column"] = image_columns[-1]
                best_mapping = current_mapping

        result = best_mapping
        log_event("header_hints_detected", hints_found=len(result), match_count=best_match_count)
        return result


def build_semantic_cells(row_values: tuple[Any, ...]) -> list[dict[str, Any]]:
    cells: list[dict[str, Any]] = []

    for column_index, cell_value in enumerate(row_values, start=1):
        text = normalize_text(cell_value)
        if not text:
            continue

        cells.append(
            {
                "column": column_index,
                "text": text,
                "normalized": normalize_key(text),
                "header_key": normalize_header(text),
            }
        )

    return cells


def build_row_snapshot(cells: list[dict[str, Any]], row_images: dict[str, Any]) -> dict[str, Any]:
    return {
        "cells": [{"column": cell["column"], "value": cell["text"]} for cell in cells],
        "images_present": sorted(row_images.keys()),
    }


def is_separator_row(cells: list[dict[str, Any]]) -> bool:
    if not cells:
        return False
    return all(not any(ch.isalnum() for ch in cell["text"]) for cell in cells)


def is_header_row(cells: list[dict[str, Any]]) -> bool:
    matched_fields = {match_header_field(cell["text"]) for cell in cells}
    matched_fields.discard(None)
    return len(matched_fields) >= 3


def looks_like_op_number(value: str) -> bool:
    if not value:
        return False
    return bool(OP_NUMBER_PATTERN.match(value.strip()))


def parse_op_value(value: Any) -> str | None:
    text = normalize_text(value)
    if not text:
        return None

    if looks_like_op_number(text):
        match = re.fullmatch(r"OP\.?\s*(\d+)([A-Z]*)", text.strip(), re.IGNORECASE)
        if not match:
            return None
        digits, suffix = match.groups()
        return f"OP {digits}{suffix.upper()}"

    if re.fullmatch(r"\d+(?:\.0+)?", text):
        numeric_text = text.split(".", 1)[0]
        return f"OP {numeric_text}"

    return None


def looks_like_fixture_number(value: str) -> bool:
    candidate = normalize_fixture_number(value)
    if not candidate:
        return False
    if looks_like_op_number(candidate):
        return False
    if "SCOPE" in candidate:
        return False
    if candidate.startswith("WBS-"):
        return False
    return bool(FIXTURE_NUMBER_PATTERN.match(candidate))


def parse_qty_value(value: Any) -> int | None:
    text = normalize_text(value)
    if not text:
        return None
    if re.fullmatch(r"\d+", text):
        qty = int(text)
        return qty if qty > 0 else None
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        qty = int(float(text))
        return qty if qty > 0 else None
    return None


def looks_like_fixture_type(value: str) -> bool:
    normalized = normalize_key(value)
    if not normalized:
        return False
    return any(keyword in normalized for keyword in FIXTURE_TYPE_KEYWORDS)



def choose_hint_text(cell_map: dict[int, str], column_index: int | None, validator=None) -> str:
    if not column_index:
        return ""
    text = cell_map.get(column_index, "")
    if not text:
        return ""
    if validator and not validator(text):
        return ""
    return text


def is_valid_field_value(field_name: str, value: str) -> bool:
    if not value:
        return False
    if field_name == "row_reference":
        return True
    if field_name == "fixture_no":
        return looks_like_fixture_number(value)
    if field_name == "op_no":
        return parse_op_value(value) is not None
    if field_name == "qty":
        return parse_qty_value(value) is not None
    if field_name == "fixture_type":
        return looks_like_fixture_type(value)
    if field_name == "part_name":
        return True
    return False


def should_carry_field_value(field_name: str) -> bool:
    return field_name in {"fixture_no", "fixture_type", "qty"}


def normalize_business_row_reference(value: Any) -> str | None:
    text = normalize_text(value)
    if not text:
        return None
    return text


def looks_like_business_row_reference(value: Any) -> bool:
    text = normalize_text(value)
    if not text:
        return False

    if looks_like_fixture_number(text) or looks_like_op_number(text):
        return False

    return bool(re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9._/-]*", text))


def choose_single_candidate(candidates: list[dict[str, Any]], field_label: str, row_index: int, snapshot: dict[str, Any]):
    if len(candidates) == 1:
        return candidates[0], None
    if not candidates:
        return None, build_error(
            f"Could not confidently extract required field: {field_label}.",
            excel_row=row_index,
            raw_data={**snapshot, "candidate_field": field_label},
        )
    return None, build_error(
        f"Multiple possible values found for {field_label}; row rejected to avoid guessing.",
        excel_row=row_index,
        raw_data={
            **snapshot,
            "candidate_field": field_label,
            "candidate_values": [{"column": cell["column"], "value": cell["text"]} for cell in candidates],
        },
    )


def parse_fixture_candidate_optimized(
    row_index: int,
    cells: list[dict[str, Any]],
    header_hints: dict[str, int],
    row_images: dict[str, Any],
    *,
    sheet_name: str,
    inherited_hints: dict[str, str] | None = None,
    fixture_pattern: re.Pattern,
    op_pattern: re.Pattern,
    qty_pattern: re.Pattern,
) -> tuple[dict[str, Any] | None, dict[str, Any] | None]:
    inherited_hints = inherited_hints or {}
    snapshot = build_row_snapshot(cells, row_images)
    snapshot["sheet_name"] = sheet_name
    if inherited_hints:
        snapshot["inherited_hints"] = inherited_hints
    cell_map = {cell["column"]: cell["text"] for cell in cells}
    cells_by_column = {cell["column"]: cell for cell in cells}
    used_columns: set[int] = set()
    parsed = {
        "fixture_no": "",
        "op_no": "",
        "part_name": "",
        "fixture_type": "",
        "qty": "",
    }
    if header_hints.get("remark"):
        used_columns.add(header_hints["remark"])

    if all(field_name in header_hints for field_name in REQUIRED_HEADER_FIELDS):
        direct_values = {
            field_name: normalize_text(cell_map.get(header_hints[field_name], ""))
            for field_name in REQUIRED_HEADER_FIELDS
        }
        has_required_column_data = any(direct_values.values()) or bool(row_images)
        if not has_required_column_data:
            return None, None

        parsed["fixture_no"] = normalize_fixture_number(direct_values["fixture_no"])
        parsed["op_no"] = parse_op_value(direct_values["op_no"]) or ""
        parsed["part_name"] = direct_values["part_name"]
        parsed["fixture_type"] = direct_values["fixture_type"]
        qty_value = parse_qty_value(direct_values["qty"])
        parsed["qty"] = str(qty_value) if qty_value is not None else ""

        validation_errors: list[tuple[str, str, str, str]] = []
        if not direct_values["fixture_no"]:
            validation_errors.append(("Fixture No", "Missing Fixture No.", direct_values["fixture_no"], "A PARC fixture number such as PARC25119001"))
        elif not looks_like_fixture_number(direct_values["fixture_no"]):
            validation_errors.append(("Fixture No", "Invalid Fixture No format.", direct_values["fixture_no"], "A PARC fixture number such as PARC25119001"))

        if not direct_values["op_no"]:
            validation_errors.append(("OP.NO", "Missing OP.NO.", direct_values["op_no"], "OP format such as OP 10"))
        elif not parsed["op_no"]:
            validation_errors.append(("OP.NO", "Invalid OP.NO format.", direct_values["op_no"], "OP format such as OP 10"))

        if not direct_values["part_name"]:
            validation_errors.append(("Part Name", "Missing Part Name.", direct_values["part_name"], "Non-empty part name"))

        if not direct_values["fixture_type"]:
            validation_errors.append(("Fixture Type", "Missing Fixture Type.", direct_values["fixture_type"], "Non-empty fixture type"))

        if not direct_values["qty"]:
            validation_errors.append(("QTY", "Missing QTY.", direct_values["qty"], "A positive number such as 1"))
        elif qty_value is None:
            validation_errors.append(("QTY", "Invalid QTY format.", direct_values["qty"], "A positive number such as 1"))

        if validation_errors:
            field_name, message, detected_value, expected_value = validation_errors[0]
            return None, build_error(
                message,
                excel_row=row_index,
                raw_data={
                    **snapshot,
                    "parsed": parsed,
                    "candidate_field": field_name,
                    "detected_value": detected_value,
                    "expected": expected_value,
                },
            )

        hint_row_reference_column = header_hints.get("row_reference")
        business_row_reference = normalize_business_row_reference(
            choose_hint_text(cell_map, hint_row_reference_column)
        )
        row_reference = business_row_reference or str(row_index)
        row_reference_source = "business_serial" if business_row_reference else "excel_row"
        row_number = int(business_row_reference) if business_row_reference and re.fullmatch(r"\d+", business_row_reference) else row_index

        return (
            {
                "excel_row": row_index,
                "row_number": row_number,
                "row_reference": row_reference,
                "row_reference_source": row_reference_source,
                "business_row_reference": business_row_reference,
                "fixture_no": parsed["fixture_no"],
                "op_no": parsed["op_no"],
                "part_name": parsed["part_name"],
                "fixture_type": parsed["fixture_type"],
                "qty": parsed["qty"],
                "remark": None,
                "image_1_url": None,
                "image_2_url": None,
                "image_1_upload": row_images.get("image_1_url"),
                "image_2_upload": None,
                "parser_confidence": "HIGH",
                "raw_data": {
                    **snapshot,
                    "excel_row": row_index,
                    "row_reference": row_reference,
                    "row_reference_source": row_reference_source,
                    "business_row_reference": business_row_reference,
                    "normalized_fields": {
                        "fixture_no": parsed["fixture_no"] or None,
                        "op_no": parsed["op_no"] or None,
                        "part_name": parsed["part_name"] or None,
                        "fixture_type": parsed["fixture_type"] or None,
                        "qty": parsed["qty"] or None,
                        "image_1_present": bool(row_images.get("image_1_url")),
                        "image_2_url": None,
                    },
                },
            },
            None,
        )

    # Performance: Use pre-compiled patterns for faster matching
    fixture_candidates = [cell for cell in cells if fixture_pattern.match(cell["text"])]
    if not fixture_candidates:
        if row_images:
            return None, build_error(
                "Row contains mapped images but no valid fixture number pattern.",
                excel_row=row_index,
                raw_data=snapshot,
            )
        return None, None

    fixture_cell, fixture_error = choose_single_candidate(fixture_candidates, "FIXTURE NO", row_index, snapshot)
    if fixture_error:
        return None, fixture_error

    parsed["fixture_no"] = normalize_fixture_number(fixture_cell["text"])
    used_columns.add(fixture_cell["column"])

    hint_row_reference_column = header_hints.get("row_reference")
    business_row_reference = normalize_business_row_reference(
        choose_hint_text(cell_map, hint_row_reference_column)
    )
    if business_row_reference and hint_row_reference_column:
        used_columns.add(hint_row_reference_column)
    else:
        row_reference_candidates = [
            cell
            for cell in cells
            if cell["column"] not in used_columns
            and cell["column"] < fixture_cell["column"]
            and looks_like_business_row_reference(cell["text"])
        ]
        if len(row_reference_candidates) == 1:
            business_row_reference = normalize_business_row_reference(row_reference_candidates[0]["text"])
            used_columns.add(row_reference_candidates[0]["column"])

    hint_op_no = choose_hint_text(cell_map, header_hints.get("op_no"), lambda value: parse_op_value(value) is not None)
    if hint_op_no:
        parsed["op_no"] = parse_op_value(hint_op_no) or normalize_text(hint_op_no)
        if "op_no" in header_hints:
            used_columns.add(header_hints["op_no"])
    else:
        # Performance: Use pre-compiled pattern
        op_candidates = [cell for cell in cells if cell["column"] not in used_columns and op_pattern.match(cell["text"])]
        if op_candidates:
            op_cell, op_error = choose_single_candidate(op_candidates, "OP.NO", row_index, snapshot)
            if op_error:
                return None, op_error
            parsed["op_no"] = parse_op_value(op_cell["text"]) or normalize_text(op_cell["text"])
            used_columns.add(op_cell["column"])
        elif inherited_hints.get("op_no"):
            parsed["op_no"] = parse_op_value(inherited_hints["op_no"]) or normalize_text(inherited_hints["op_no"])

    hint_qty_column = header_hints.get("qty")
    hint_qty_value = parse_qty_value(choose_hint_text(cell_map, hint_qty_column))
    if hint_qty_value is not None:
        parsed["qty"] = str(hint_qty_value)
        if hint_qty_column:
            used_columns.add(hint_qty_column)
    else:
        # Performance: Use pre-compiled pattern
        qty_candidates = [cell for cell in cells if cell["column"] not in used_columns and qty_pattern.match(cell["text"])]
        if qty_candidates:
            qty_cell, qty_error = choose_single_candidate(qty_candidates, "QTY", row_index, snapshot)
            if qty_error:
                return None, qty_error
            parsed["qty"] = str(parse_qty_value(qty_cell["text"]))
            used_columns.add(qty_cell["column"])
        elif inherited_hints.get("qty"):
            inherited_qty = parse_qty_value(inherited_hints.get("qty"))
            if inherited_qty is not None:
                parsed["qty"] = str(inherited_qty)

    hint_fixture_type_column = header_hints.get("fixture_type")
    hint_fixture_type = choose_hint_text(cell_map, hint_fixture_type_column)
    if hint_fixture_type:
        parsed["fixture_type"] = hint_fixture_type
        if hint_fixture_type_column:
            used_columns.add(hint_fixture_type_column)
    else:
        fixture_type_candidates = [
            cell for cell in cells if cell["column"] not in used_columns and looks_like_fixture_type(cell["text"])
        ]
        fixture_type_cell, fixture_type_error = choose_single_candidate(
            fixture_type_candidates,
            "Fixture Type",
            row_index,
            snapshot,
        )
        if fixture_type_error:
            return None, fixture_type_error
        if fixture_type_cell:
            parsed["fixture_type"] = fixture_type_cell["text"]
            used_columns.add(fixture_type_cell["column"])
        elif inherited_hints.get("fixture_type"):
            parsed["fixture_type"] = inherited_hints["fixture_type"]

    hint_part_name_column = header_hints.get("part_name")
    hint_part_name = choose_hint_text(cell_map, hint_part_name_column)
    if hint_part_name:
        parsed["part_name"] = hint_part_name
        if hint_part_name_column:
            used_columns.add(hint_part_name_column)
    else:
        part_name_candidates = []
        for cell in cells:
            if cell["column"] in used_columns:
                continue
            if match_header_field(cell["text"]):
                continue
            if looks_like_fixture_type(cell["text"]):
                continue
            if op_pattern.match(cell["text"]):
                continue
            if fixture_pattern.match(cell["text"]):
                continue
            if qty_pattern.match(cell["text"]):
                continue
            part_name_candidates.append(cell)

        part_name_cell = None
        if len(part_name_candidates) > 1:
            return None, build_error(
                "Multiple possible values found for Part Name; row rejected to avoid guessing.",
                excel_row=row_index,
                raw_data={
                    **snapshot,
                    "candidate_field": "Part Name",
                    "candidate_values": [{"column": cell["column"], "value": cell["text"]} for cell in part_name_candidates],
                },
            )
        if len(part_name_candidates) == 1:
            part_name_cell = part_name_candidates[0]

        if part_name_cell:
            parsed["part_name"] = part_name_cell["text"]
            used_columns.add(part_name_cell["column"])
        else:
            merged_part_name_source = choose_hint_text(cell_map, header_hints.get("op_no"))
            if (
                merged_part_name_source
                and not parse_op_value(merged_part_name_source)
                and not fixture_pattern.match(merged_part_name_source)
                and parse_qty_value(merged_part_name_source) is None
            ):
                parsed["part_name"] = merged_part_name_source
        if not parsed["part_name"] and inherited_hints.get("part_name"):
            parsed["part_name"] = inherited_hints["part_name"]

    missing_fields = [
        label
        for field_name, label in (
            ("fixture_no", "FIXTURE NO"),
            ("part_name", "Part Name"),
            ("fixture_type", "Fixture Type"),
            ("qty", "QTY"),
        )
        if not parsed[field_name]
    ]

    if missing_fields:
        return None, build_error(
            f"Could not confidently extract required fields: {', '.join(missing_fields)}.",
            excel_row=row_index,
            raw_data={**snapshot, "parsed": parsed},
        )

    row_reference = business_row_reference or str(row_index)
    row_reference_source = "business_serial" if business_row_reference else "excel_row"
    row_number = int(business_row_reference) if business_row_reference and re.fullmatch(r"\d+", business_row_reference) else row_index

    return (
        {
            "excel_row": row_index,
            "row_number": row_number,
            "row_reference": row_reference,
            "row_reference_source": row_reference_source,
            "business_row_reference": business_row_reference,
            "fixture_no": parsed["fixture_no"],
            "op_no": parsed["op_no"],
            "part_name": parsed["part_name"],
            "fixture_type": parsed["fixture_type"],
            "qty": parsed["qty"],
            "remark": None,
            "image_1_url": None,
            "image_2_url": None,
            "image_1_upload": row_images.get("image_1_url"),
            "image_2_upload": None,
            "parser_confidence": "HIGH",
            "raw_data": {
                **snapshot,
                "excel_row": row_index,
                "row_reference": row_reference,
                "row_reference_source": row_reference_source,
                "business_row_reference": business_row_reference,
                "normalized_fields": {
                    "fixture_no": parsed["fixture_no"] or None,
                    "op_no": parsed["op_no"] or None,
                    "part_name": parsed["part_name"] or None,
                    "fixture_type": parsed["fixture_type"] or None,
                    "qty": parsed["qty"] or None,
                    "remark": None,
                    "image_1_present": bool(row_images.get("image_1_url")),
                    "image_2_url": None,
                },
            },
        },
        None,
    )


def build_vertical_merge_lookup(worksheet) -> dict[tuple[int, int], Any]:
    lookup: dict[tuple[int, int], Any] = {}

    for merged_range in getattr(worksheet.merged_cells, "ranges", []):
        if merged_range.min_col != merged_range.max_col:
            continue

        master_value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
        for row_index in range(merged_range.min_row, merged_range.max_row + 1):
            lookup[(row_index, merged_range.min_col)] = master_value

    return lookup


def get_effective_row_values_fast(worksheet, row_index: int, max_column: int, vertical_merge_lookup: dict[tuple[int, int], Any]) -> tuple[Any, ...]:
    """Optimized row value extraction using iter_rows for better performance."""
    # Use iter_rows with values_only for much faster bulk cell access
    row_values = next(worksheet.iter_rows(min_row=row_index, max_row=row_index, max_col=max_column, values_only=True), ())

    # Apply merge lookup only for cells that are empty
    if vertical_merge_lookup:
        values: list[Any] = []
        for col_idx, cell_value in enumerate(row_values, start=1):
            if cell_value in (None, ""):
                cell_value = vertical_merge_lookup.get((row_index, col_idx), cell_value)
            values.append(cell_value)
        # Pad with None if row is shorter than max_column
        while len(values) < max_column:
            values.append(vertical_merge_lookup.get((row_index, len(values) + 1), None))
        return tuple(values)

    # Pad with None if row is shorter than max_column (no merge lookup needed)
    values = list(row_values)
    while len(values) < max_column:
        values.append(None)
    return tuple(values)


def build_rows(
    worksheet,
    metadata_row: int,
    header_hints: dict[str, int],
    images_by_row: dict[int, dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    missing_headers = missing_required_headers(header_hints)

    if missing_headers:
        return [], [
            build_error(
                f"Missing required Excel headers: {', '.join(missing_headers)}.",
                excel_row=metadata_row if metadata_row > 0 else None,
                raw_data={
                    "reason": "missing_required_headers",
                    "missing_fields": missing_headers,
                    "expected_headers": EXPECTED_HEADER_LABELS,
                },
            )
        ]

    # Performance: Build merge lookup only if needed
    merge_lookup_start = time.perf_counter()
    vertical_merge_lookup = build_vertical_merge_lookup(worksheet) if worksheet.merged_cells.ranges else {}
    merge_lookup_time = time.perf_counter() - merge_lookup_start
    log_event("merge_lookup_timing", 
             sheet_title=worksheet.title,
             time_ms=round(merge_lookup_time * 1000, 2),
             merged_ranges=len(worksheet.merged_cells.ranges))

    carry_hints: dict[str, str] = {}
    header_row = int(header_hints.get("__row") or metadata_row)
    start_row = max(1, metadata_row + 1, header_row + 1)
    max_row = worksheet.max_row
    max_column = worksheet.max_column

    log_event("build_rows_start", start_row=start_row, max_row=max_row, max_column=max_column)

    # Performance: Pre-compile regex patterns for faster matching
    fixture_pattern = re.compile(r"^PARC\d{4,}$", re.IGNORECASE)
    op_pattern = re.compile(r"^OP\.?\s*\d+[A-Z]*$", re.IGNORECASE)
    qty_pattern = re.compile(r"^\d+(?:\.0+)?$")
    
    # Performance: Batch process rows for better memory usage
    row_count = 0
    processed_rows = 0
    
    with Timer("iterate_rows"):
        # Use values_only for much faster iteration
        for row_index, row_values in enumerate(
            worksheet.iter_rows(min_row=start_row, max_row=max_row, max_col=max_column, values_only=True),
            start=start_row
        ):
            row_count += 1
            row_images = images_by_row.get(row_index, {})

            # Performance: Skip empty rows early
            if not any(row_values) and not row_images:
                continue

            # Performance: Apply merge lookup efficiently
            if vertical_merge_lookup and row_values:
                effective_values = list(row_values)
                # Only check for merged values in empty cells
                for col_idx, cell_value in enumerate(row_values, start=1):
                    if cell_value in (None, "") and (row_index, col_idx) in vertical_merge_lookup:
                        effective_values[col_idx - 1] = vertical_merge_lookup[(row_index, col_idx)]
                # Pad if necessary
                while len(effective_values) < max_column:
                    col_idx = len(effective_values) + 1
                    if (row_index, col_idx) in vertical_merge_lookup:
                        effective_values.append(vertical_merge_lookup[(row_index, col_idx)])
                    else:
                        effective_values.append(None)
                row_values = tuple(effective_values)

            # Performance: Fast cell building with early filtering
            cells = []
            for column_index, cell_value in enumerate(row_values, start=1):
                text = normalize_text(cell_value)
                if not text:
                    continue
                
                # Quick filter for common non-data content
                if text.startswith("WBS-") or text.upper().startswith(("TOTAL", "SUM", "SUBTOTAL")):
                    continue
                    
                cells.append({
                    "column": column_index,
                    "text": text,
                    "normalized": normalize_key(text),
                    "header_key": normalize_header(text),
                })

            if not cells and not row_images:
                continue

            # Performance: Fast row type detection
            if not cells:
                # Only images case - already handled above
                pass
            elif is_separator_row(cells):
                continue
            elif is_header_row(cells):
                continue

            processed_rows += 1
            row_has_structured_content = bool(cells or row_images)
            inherited_hints: dict[str, str] = {}
            
            # Performance: Optimized hint processing
            for field_name, column_index in header_hints.items():
                if field_name.startswith("__"):
                    continue
                if not column_index:
                    continue

                raw_value = ""
                if column_index - 1 < len(row_values):
                    raw_value = normalize_text(row_values[column_index - 1])
                
                if raw_value and is_valid_field_value(field_name, raw_value):
                    carry_hints[field_name] = raw_value
                    continue

                if row_has_structured_content and field_name in carry_hints and should_carry_field_value(field_name):
                    inherited_hints[field_name] = carry_hints[field_name]

            # Performance: Parse fixture candidate with optimized patterns
            parsed_row, error = parse_fixture_candidate_optimized(
                row_index,
                cells,
                header_hints,
                row_images,
                sheet_name=worksheet.title,
                inherited_hints=inherited_hints,
                fixture_pattern=fixture_pattern,
                op_pattern=op_pattern,
                qty_pattern=qty_pattern,
            )

            if error:
                errors.append(error)
                continue

            if parsed_row:
                # Performance: Update carry hints efficiently
                for field_name in ("fixture_no", "op_no", "part_name", "fixture_type", "qty"):
                    value = parsed_row.get(field_name)
                    if value:
                        carry_hints[field_name] = str(value)
                rows.append(parsed_row)
                continue

            if row_images:
                errors.append(
                    build_error(
                        "Row contains mapped images but no valid fixture data.",
                        excel_row=row_index,
                        raw_data=build_row_snapshot(cells, row_images),
                    )
                )

    log_event("build_rows_complete", 
             rows_scanned=row_count, 
             rows_processed=processed_rows,
             rows_accepted=len(rows), 
             errors_count=len(errors),
             processing_rate=round(processed_rows / max(time.perf_counter() - merge_lookup_start, 0.001), 2))
    return rows, errors


def _process_workbook(file_bytes: bytes) -> dict[str, Any]:
    start_time = time.perf_counter()
    log_event("process_workbook_start", file_size_bytes=len(file_bytes))
    
    with Timer("workbook_load"):
        workbook = open_excel_workbook(file_bytes, read_only=False)
    
    try:
        with Timer("find_metadata"):
            _metadata_sheet_name, _metadata_row, metadata_value = find_workbook_metadata(workbook)
        
        with Timer("parse_wbs_header"):
            file_info = parse_wbs_header(metadata_value)

        all_rows: list[dict[str, Any]] = []
        all_errors: list[dict[str, Any]] = []
        
        # Performance tracking per worksheet
        for worksheet_idx, worksheet in enumerate(workbook.worksheets):
            worksheet_start = time.perf_counter()
            log_event("processing_worksheet", 
                     worksheet_index=worksheet_idx,
                     sheet_title=worksheet.title, 
                     max_row=worksheet.max_row, 
                     max_column=worksheet.max_column)
            
            try:
                with Timer(f"find_metadata_{worksheet.title}"):
                    metadata_row, _ = find_metadata_row(worksheet)
            except ValueError:
                metadata_row = 0

            # Performance: Header detection optimization
            with Timer(f"detect_header_hints_{worksheet.title}"):
                header_hints = detect_header_hints(worksheet, metadata_row)

            if metadata_row == 0 and missing_required_headers(header_hints):
                log_event("worksheet_skipped_no_wbs_or_strict_headers", sheet_title=worksheet.title)
                continue
            
            # Performance: Image extraction with detailed timing
            image_extraction_start = time.perf_counter()
            with Timer(f"extract_images_{worksheet.title}"):
                images_by_row, image_errors = extract_anchored_images(worksheet, header_hints)
            image_extraction_time = time.perf_counter() - image_extraction_start
            log_event("image_extraction_timing",
                     sheet_title=worksheet.title,
                     extraction_time_ms=round(image_extraction_time * 1000, 2),
                     images_found=len(images_by_row),
                     image_errors=len(image_errors))
            
            # Performance: Row building with detailed timing
            row_building_start = time.perf_counter()
            with Timer(f"build_rows_{worksheet.title}"):
                rows, parsing_errors = build_rows(worksheet, metadata_row, header_hints, images_by_row)
            row_building_time = time.perf_counter() - row_building_start
            
            worksheet_time = time.perf_counter() - worksheet_start
            log_event("worksheet_processing_complete",
                     worksheet_index=worksheet_idx,
                     sheet_title=worksheet.title,
                     total_time_ms=round(worksheet_time * 1000, 2),
                     row_building_time_ms=round(row_building_time * 1000, 2),
                     rows_processed=len(rows),
                     errors_count=len(parsing_errors),
                     images_mapped=len(images_by_row))

            all_rows.extend(rows)
            all_errors.extend(image_errors)
            all_errors.extend(parsing_errors)
    finally:
        workbook.close()

    # Performance: Deduplication timing
    with Timer("deduplication"):
        deduped_rows: list[dict[str, Any]] = []
        seen_row_keys: set[tuple[str, int, str]] = set()
        for row in all_rows:
            row_key = (
                row.get("raw_data", {}).get("sheet_name", ""),
                int(row["excel_row"]),
                row["fixture_no"],
            )
            if row_key in seen_row_keys:
                continue
            seen_row_keys.add(row_key)
            deduped_rows.append(row)

    # Performance: Final validation and response building
    with Timer("final_validation"):
        errors = all_errors
        if not deduped_rows and not errors:
            errors.append(build_error("No fixture rows were found in the workbook."))

        for row in deduped_rows:
            log_event(
                "extract_row_accepted",
                excel_row=row["excel_row"],
                fixture_no=row["fixture_no"],
                parser_confidence=row["parser_confidence"],
            )

        for error in errors:
            log_event(
                "extract_row_rejected",
                excel_row=error.get("excel_row"),
                reason=error.get("error_message"),
                raw_data=error.get("raw_data", {}),
            )
    
    total_processing_time = time.perf_counter() - start_time
    log_event("process_workbook_complete", 
             total_time_ms=round(total_processing_time * 1000, 2),
             final_rows_count=len(deduped_rows),
             final_errors_count=len(errors),
             avg_time_per_row=round(total_processing_time * 1000 / max(len(deduped_rows), 1), 2))

    return {
        "file_info": {
            "project_code": file_info["project_code"],
            "project_name": file_info["project_name"],
            "company_name": file_info["company_name"],
        },
        "rows": deduped_rows,
        "errors": errors,
    }


@app.post("/extract")
async def extract_workbook(
    request: Request,
    x_extraction_token: str | None = Header(default=None),
):
    content_type = normalize_text(request.headers.get("content-type"))

    if not SERVICE_TOKEN:
        log_event("extract_request_rejected", reason="service_not_configured", content_type=content_type)
        return build_error_response(500, "Service is not configured", [build_error("EXTRACTION_SERVICE_TOKEN is required.")])

    form = await request.form()
    form_keys = list(form.keys())
    upload_fields = []
    for field_name, value in form.multi_items():
        if isinstance(value, StarletteUploadFile):
            upload_fields.append(
                {
                    "field_name": field_name,
                    "filename": normalize_text(value.filename),
                    "content_type": normalize_text(value.content_type),
                }
            )

    log_event(
        "extract_request_received",
        content_type=content_type,
        form_keys=form_keys,
        upload_fields=upload_fields,
    )

    if x_extraction_token != SERVICE_TOKEN:
        log_event("extract_request_rejected", reason="invalid_token", content_type=content_type, form_keys=form_keys)
        return build_error_response(401, "Unauthorized", [build_error("Invalid extraction token.")])

    if "multipart/form-data" not in content_type.lower():
        log_event("extract_request_rejected", reason="invalid_content_type", content_type=content_type, form_keys=form_keys)
        return build_error_response(422, "Failed to process file", [build_error("Request must use multipart/form-data.")])

    if EXPECTED_UPLOAD_FIELD not in form:
        log_event(
            "extract_request_rejected",
            reason="missing_expected_upload_field",
            expected_field=EXPECTED_UPLOAD_FIELD,
            form_keys=form_keys,
            upload_fields=upload_fields,
        )
        return build_error_response(
            422,
            "Failed to process file",
            [build_error(f"Expected multipart field '{EXPECTED_UPLOAD_FIELD}' was not provided.")],
        )

    uploaded_file = form.get(EXPECTED_UPLOAD_FIELD)
    if not isinstance(uploaded_file, StarletteUploadFile):
        log_event(
            "extract_request_rejected",
            reason="invalid_upload_field_type",
            expected_field=EXPECTED_UPLOAD_FIELD,
            form_keys=form_keys,
        )
        return build_error_response(
            422,
            "Failed to process file",
            [build_error(f"Multipart field '{EXPECTED_UPLOAD_FIELD}' must contain a file upload.")],
        )

    if len(upload_fields) != 1:
        log_event(
            "extract_request_rejected",
            reason="unexpected_upload_field_count",
            upload_fields=upload_fields,
        )
        return build_error_response(
            422,
            "Failed to process file",
            [build_error("Exactly one uploaded Excel file is required.")],
        )

    if not normalize_text(uploaded_file.filename).lower().endswith(ALLOWED_EXTENSION):
        log_event(
            "extract_request_rejected",
            reason="invalid_extension",
            filename=normalize_text(uploaded_file.filename),
        )
        return build_error_response(400, "Failed to process file", [build_error("Only .xlsx files are allowed.")])

    if normalize_text(uploaded_file.content_type).lower() != ALLOWED_CONTENT_TYPE:
        log_event(
            "extract_request_rejected",
            reason="invalid_mime_type",
            filename=normalize_text(uploaded_file.filename),
            file_content_type=normalize_text(uploaded_file.content_type),
        )
        return build_error_response(400, "Failed to process file", [build_error("Only .xlsx MIME type is allowed.")])

    file_bytes = await uploaded_file.read()
    if not file_bytes:
        log_event("extract_request_rejected", reason="empty_file", filename=normalize_text(uploaded_file.filename))
        return build_error_response(400, "Failed to process file", [build_error("Uploaded file is empty.")])

    if len(file_bytes) > MAX_UPLOAD_BYTES:
        log_event(
            "extract_request_rejected",
            reason="file_too_large",
            filename=normalize_text(uploaded_file.filename),
            size_bytes=len(file_bytes),
        )
        return build_error_response(400, "Failed to process file", [build_error("Excel file exceeds the maximum allowed size.")])

    log_event(
        "extract_processing_started",
        filename=normalize_text(uploaded_file.filename),
        size_bytes=len(file_bytes),
    )

    try:
        with Timer("total_processing"):
            result = await asyncio.to_thread(_process_workbook, file_bytes)
        log_event(
            "extract_processing_completed",
            filename=normalize_text(uploaded_file.filename),
            accepted_rows=len(result["rows"]),
            error_count=len(result["errors"]),
        )
        return result
    except ValueError as exc:
        log_event("extract_processing_failed", filename=normalize_text(uploaded_file.filename), reason=str(exc))
        return build_error_response(422, "Failed to process file", [build_error(str(exc))])
    except Exception as exc:
        log_event("extract_processing_failed", filename=normalize_text(uploaded_file.filename), reason=str(exc))
        return build_error_response(500, "Failed to process file", [build_error(str(exc))])


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(
        "app.main:app",
        host=os.getenv("HOST", "0.0.0.0"),
        port=int(os.getenv("PORT", "8000")),
    )
