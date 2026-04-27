"""
format_scope_report.py
Usage: python format_scope_report.py <raw_input.xlsx> <formatted_output.xlsx>

Reads the raw JS-generated Excel file and produces the final formatted report.
This script is the ONLY styling authority. The JS layer must not apply any styling.

Expected raw sheet columns (row 1 = WBS title, row 2 = headers, row 3+ = data):
  A  (col 1):  S. No
  B  (col 2):  FIXTURE NO
  C  (col 3):  OP.NO
  D  (col 4):  Part Name
  E  (col 5):  Fixture Type
  F  (col 6):  QTY
  G  (col 7):  Status
  H  (col 8):  CONCEPT
  I  (col 9):  CONCEPT hrs
  J  (col 10): DAP
  K  (col 11): DAP hrs
  L  (col 12): 3D FINISH
  M  (col 13): 3D FINISH hrs
  N  (col 14): 2D FINISH
  O  (col 15): 2D FINISH hrs
  P  (col 16): Total hrs
  Q  (col 17): Designer
  R  (col 18): Ref Image      ← hyperlink to image_1_url
  S  (col 19): Work Image     ← hyperlink to image_2_url
  T  (col 20): Work Proof     ← hyperlink to proof url

Output sheet has 19 columns (Work Image merged visually into Ref Image column,
both hyperlinks preserved; only one "Ref. Image" column shown in output).
"""

from __future__ import annotations

import sys
from collections import Counter

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Colour palette  (all 6-char hex, no alpha prefix)
# ---------------------------------------------------------------------------
C_TITLE_BG       = "0D1B3E"
C_TITLE_FG       = "FFFFFF"
C_SECT_LEFT_BG   = "1A3A6C"
C_SECT_MID_BG    = "163561"
C_SECT_RIGHT_BG  = "1A3A6C"
C_SECT_FG        = "FFFFFF"
C_HDR_FIXTURE    = "2C5F9E"
C_HDR_STAGE      = "1F4F8F"
C_HDR_SUMMARY    = "2C5F9E"
C_HDR_FG         = "FFFFFF"
C_HDR_BORDER_T   = "4A7DB5"
C_HDR_BORDER_B   = "FFFFFF"
C_DATA_BORDER    = "BDC9DD"
C_CELL_FG        = "1A1A1A"
C_SERIAL_FG      = "666666"
C_FIXTURE_FG     = "0D1B3E"
C_STAGE_EMPTY_BG = "E8EDF5"
C_TOTAL_BG       = "0D1B3E"
C_TOTAL_FG       = "FFFFFF"

# Stage state → (background, text_is_dark)
STAGE_COLORS: dict[str, tuple[str, bool]] = {
    "DONE":        ("4CAF50", False),   # green  → white text
    "IN_PROGRESS": ("FF9800", True),    # amber  → dark text
    "EMPTY":       (C_STAGE_EMPTY_BG, True),  # light blue-grey → dark text
}

# Display status → (background, text_is_dark)
STATUS_COLORS: dict[str, tuple[str, bool]] = {
    "Assigned":    ("3A7BD5", False),
    "In Progress": ("28A745", False),
    "On Hold":     ("FF9800", False),
    "Review":      ("009688", False),
    "Rework":      ("9C27B0", False),
    "Closed":      ("616161", False),
    "Overdue":     ("D32F2F", False),
}

# Raw status value from JS → display label
RAW_STATUS_MAP: dict[str, str] = {
    "ASSIGNED":    "Assigned",
    "IN_PROGRESS": "In Progress",
    "HOLD":        "On Hold",
    "ON_HOLD":     "On Hold",
    "REVIEW":      "Review",
    "UNDER_REVIEW": "Review",
    "REWORK":      "Rework",
    "CLOSED":      "Closed",
    "OVERDUE":     "Overdue",
    "COMPLETE":    "Closed",   # refined per-row if a stage is still TBD
    "COMPLETED":   "Closed",
}

# ---------------------------------------------------------------------------
# Output column indices  (1-based)
# ---------------------------------------------------------------------------
COL_SERIAL   = 1
COL_FIXTURE  = 2
COL_OPNO     = 3
COL_PART     = 4
COL_FXTYPE   = 5
COL_QTY      = 6
COL_STATUS   = 7
COL_CONCEPT  = 8
COL_CONHRS   = 9
COL_DAP      = 10
COL_DAPHRS   = 11
COL_3D       = 12
COL_3DHRS    = 13
COL_2D       = 14
COL_2DHRS    = 15
COL_TOTAL    = 16
COL_DESIGNER = 17
COL_REFIMG   = 18
COL_FIXIMG   = 19   # NEW COLUMN
COL_PROOF    = 20   # SHIFTED
OUT_NCOLS    = 20

# Fixed column widths
COL_WIDTHS: dict[int, float] = {
    COL_SERIAL:   5.5,
    COL_FIXTURE:  15.0,
    COL_OPNO:     8.0,
    COL_PART:     32.0,
    COL_FXTYPE:   26.0,
    COL_QTY:      5.5,
    COL_STATUS:   13.0,
    COL_CONCEPT:  26.0,
    COL_CONHRS:   11.0,
    COL_DAP:      26.0,
    COL_DAPHRS:   11.0,
    COL_3D:       26.0,
    COL_3DHRS:    11.0,
    COL_2D:       26.0,
    COL_2DHRS:    11.0,
    COL_TOTAL:    12.0,
    COL_DESIGNER: 11.0,
    COL_REFIMG:   9.0,
    COL_PROOF:    9.0,
}

# ---------------------------------------------------------------------------
# Style factory helpers
# ---------------------------------------------------------------------------

def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)


def _font(bold: bool = False, size: float = 9,
          color: str = "000000", name: str = "Calibri") -> Font:
    return Font(name=name, bold=bold, size=size, color=color)


def _align(h: str = "center", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrapText=wrap)


def _thin(color: str) -> Side:
    return Side(style="thin", color=color)


def _data_border() -> Border:
    s = _thin(C_DATA_BORDER)
    return Border(top=s, left=s, bottom=s, right=s)


def _header_border() -> Border:
    return Border(
        top    = _thin(C_HDR_BORDER_T),
        left   = _thin(C_HDR_BORDER_T),
        bottom = _thin(C_HDR_BORDER_B),
        right  = _thin(C_HDR_BORDER_T),
    )


# ---------------------------------------------------------------------------
# WBS title parsing
# ---------------------------------------------------------------------------

def parse_wbs_title(raw: str) -> tuple[str, str, str]:
    """
    'WBS-PARC2600M010-Fuel Tank weld Line-Belrise Industries Limited'
    → ('PARC2600M010', 'Fuel Tank weld Line', 'Belrise Industries Limited')
    """
    s = str(raw or "").strip()
    if s.startswith("WBS-"):
        s = s[4:]
    parts = s.split("-", 2)
    if len(parts) == 3:
        return parts[0].strip(), parts[1].strip(), parts[2].strip()
    return s, "", ""


def format_title(project: str, scope: str, company: str) -> str:
    return f"  PROJECT : {project}   |   SCOPE : {scope}   |   COMPANY : {company}"


# ---------------------------------------------------------------------------
# Status resolution
# ---------------------------------------------------------------------------

def resolve_status(raw_status: str, raw: dict) -> str:
    """
    Map the raw JS status string to a human-readable display label.
    For COMPLETE rows, demote to 'In Progress' if any stage date is still 'TBD'.
    """
    key     = str(raw_status or "").strip().upper().replace(" ", "_")
    display = RAW_STATUS_MAP.get(key, "Assigned")

    if display == "Closed":
        stage_dates = [raw["concept"], raw["dap"], raw["d3"], raw["d2"]]
        if any("TBD" in str(v) or "In Progress" in str(v) for v in stage_dates):
            display = "In Progress"

    return display


# ---------------------------------------------------------------------------
# Stage classification
# ---------------------------------------------------------------------------

def classify_stage(date_val: str, hrs_val: str) -> str:
    """Return 'DONE', 'IN_PROGRESS', or 'EMPTY'."""
    d = str(date_val or "").strip()
    h = str(hrs_val  or "").strip()

    if not d or d == "—":
        return "EMPTY"

    if "TBD" in d or "In Progress" in h or (d and not h):
        return "IN_PROGRESS"

    return "DONE"


# ---------------------------------------------------------------------------
# Read raw workbook
# ---------------------------------------------------------------------------

def read_raw(path: str) -> tuple[str, list[dict]]:
    wb = load_workbook(path)
    ws = wb.active

    wbs_title = str(ws.cell(1, 1).value or "")
    rows: list[dict] = []

    for r in range(3, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 21)]

        # Skip blank rows
        if not any(v for v in vals):
            continue

        def hlink(col: int) -> str | None:
            cell = ws.cell(r, col)
            return cell.hyperlink.target if cell.hyperlink else None

        rows.append({
            "serial":    vals[0],
            "fixture":   str(vals[1] or ""),
            "opno":      str(vals[2] or ""),
            "part":      str(vals[3] or ""),
            "fx_type":   str(vals[4] or ""),
            "qty":       vals[5],
            "status":    str(vals[6] or ""),
            # Stage date-range and hours (cols H-O, indices 7-14)
            "concept":   str(vals[7]  or ""),
            "con_hrs":   str(vals[8]  or ""),
            "dap":       str(vals[9]  or ""),
            "dap_hrs":   str(vals[10] or ""),
            "d3":        str(vals[11] or ""),
            "d3_hrs":    str(vals[12] or ""),
            "d2":        str(vals[13] or ""),
            "d2_hrs":    str(vals[14] or ""),
            "total_hrs": str(vals[15] or ""),
            "designer":  str(vals[16] or ""),
            # Image / proof cells — text value tells us if a link exists
            "ref_img":   str(vals[17] or ""),
            "work_img":  str(vals[18] or ""),
            "proof":     str(vals[19] or ""),
            # Actual hyperlink URLs
            "ref_url":   hlink(18),
            "work_url":  hlink(19),
            "proof_url": hlink(20),
        })

    return wbs_title, rows


# ---------------------------------------------------------------------------
# Build formatted workbook
# ---------------------------------------------------------------------------

def build_formatted(wbs_title: str, raw_rows: list[dict]) -> Workbook:
    project, scope, company = parse_wbs_title(wbs_title)
    display_title = format_title(project, scope, company)

    wb = Workbook()
    ws = wb.active
    ws.title = "Design Report"

    last_col = get_column_letter(OUT_NCOLS)

    # ── Column widths ──────────────────────────────────────────────────────
    for col_idx, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Row 1: Title bar ───────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col}1")
    c = ws.cell(1, 1, display_title)
    c.fill      = _fill(C_TITLE_BG)
    c.font      = _font(bold=True, size=13, color=C_TITLE_FG)
    c.alignment = _align(h="left", v="center")
    ws.row_dimensions[1].height = 36

    # ── Row 2: Section group headers ───────────────────────────────────────
    sections = [
        (1,  7,  "FIXTURE INFORMATION",         C_SECT_LEFT_BG),
        (8,  15, "STAGE-WISE PROGRESS TRACKING", C_SECT_MID_BG),
        (16, 20, "SUMMARY & RESOURCES",          C_SECT_RIGHT_BG),
    ]
    for start, end, label, bg in sections:
        if start != end:
            ws.merge_cells(
                start_row=2, start_column=start,
                end_row=2,   end_column=end
            )
        c = ws.cell(2, start, label)
        c.fill      = _fill(bg)
        c.font      = _font(bold=True, size=9, color=C_SECT_FG)
        c.alignment = _align()
    ws.row_dimensions[2].height = 18

    # ── Row 3: Column headers ──────────────────────────────────────────────
    headers = [
        (COL_SERIAL,   "S. No",                C_HDR_FIXTURE),
        (COL_FIXTURE,  "Fixture No",            C_HDR_FIXTURE),
        (COL_OPNO,     "OP No",                 C_HDR_FIXTURE),
        (COL_PART,     "Part Name",             C_HDR_FIXTURE),
        (COL_FXTYPE,   "Fixture Type",          C_HDR_FIXTURE),
        (COL_QTY,      "QTY",                   C_HDR_FIXTURE),
        (COL_STATUS,   "Status",                C_HDR_FIXTURE),
        (COL_CONCEPT,  "CONCEPT\nDate Range",   C_HDR_STAGE),
        (COL_CONHRS,   "CONCEPT\nhrs",          C_HDR_STAGE),
        (COL_DAP,      "DAP\nDate Range",       C_HDR_STAGE),
        (COL_DAPHRS,   "DAP\nhrs",              C_HDR_STAGE),
        (COL_3D,       "3D FINISH\nDate Range", C_HDR_STAGE),
        (COL_3DHRS,    "3D FINISH\nhrs",        C_HDR_STAGE),
        (COL_2D,       "2D FINISH\nDate Range", C_HDR_STAGE),
        (COL_2DHRS,    "2D FINISH\nhrs",        C_HDR_STAGE),
        (COL_TOTAL,    "Total\nhrs / mins",     C_HDR_SUMMARY),
        (COL_DESIGNER, "Designer",              C_HDR_SUMMARY),
        (COL_REFIMG,   "Part.\nImage",          C_HDR_SUMMARY),
        (COL_FIXIMG,   "Fix.\nImage",           C_HDR_SUMMARY),
        (COL_PROOF,    "Work\nProof",           C_HDR_SUMMARY),
    ]

    hdr_bdr = _header_border()
    for col_idx, label, bg in headers:
        c = ws.cell(3, col_idx, label)
        c.fill      = _fill(bg)
        c.font      = _font(bold=True, size=9, color=C_HDR_FG)
        c.alignment = _align(wrap=True)
        c.border    = hdr_bdr
    ws.row_dimensions[3].height = 40

    # ── Data rows ──────────────────────────────────────────────────────────
    db = _data_border()
    status_counts: Counter = Counter()
    display_row = 0

    for raw in raw_rows:
        display_row += 1
        r = 3 + display_row   # sheet row (data starts at row 4)

        display_status = resolve_status(raw["status"], raw)
        status_counts[display_status] += 1

        # Normalise stage values: empty string becomes em-dash placeholder
        def norm(v: str) -> str:
            s = v.strip()
            return s if s else "—"

        concept  = norm(raw["concept"])
        con_hrs  = norm(raw["con_hrs"])
        dap      = norm(raw["dap"])
        dap_hrs  = norm(raw["dap_hrs"])
        d3       = norm(raw["d3"])
        d3_hrs   = norm(raw["d3_hrs"])
        d2       = norm(raw["d2"])
        d2_hrs   = norm(raw["d2_hrs"])
        total    = raw["total_hrs"].strip() or None

        # Helper: write one cell with border + optional overrides
        def wc(col: int, val, font: Font = None, align: Alignment = None,
               fill: PatternFill = None) -> object:
            cell            = ws.cell(r, col, val)
            cell.border     = db
            cell.font       = font  or _font(size=9, color=C_CELL_FG)
            cell.alignment  = align or _align(h="left", v="center", wrap=True)
            if fill:
                cell.fill = fill
            return cell

        # S. No
        wc(COL_SERIAL, display_row,
           font  = _font(size=9, color=C_SERIAL_FG),
           align = _align(h="center", v="center"))

        # Fixture No — bold dark-blue
        wc(COL_FIXTURE, raw["fixture"],
           font  = _font(bold=True, size=9, color=C_FIXTURE_FG),
           align = _align(h="left", v="center"))

        # OP No
        wc(COL_OPNO, raw["opno"],
           align = _align(h="left", v="center"))

        # Part Name
        wc(COL_PART, raw["part"],
           align = _align(h="left", v="center", wrap=True))

        # Fixture Type
        wc(COL_FXTYPE, raw["fx_type"],
           align = _align(h="left", v="center", wrap=True))

        # QTY
        wc(COL_QTY, raw["qty"],
           align = _align(h="center", v="center"))

        # Status badge
        s_bg, s_dark = STATUS_COLORS.get(display_status, ("FFFFFF", True))
        s_fg = "1A1A1A" if s_dark else "FFFFFF"
        wc(COL_STATUS, display_status,
           font  = _font(bold=True, size=9, color=s_fg),
           align = _align(h="center", v="center"),
           fill  = _fill(s_bg))

        # Stage columns (date range + hrs pairs)
        stage_data = [
            (COL_CONCEPT, concept,  COL_CONHRS, con_hrs),
            (COL_DAP,     dap,      COL_DAPHRS, dap_hrs),
            (COL_3D,      d3,       COL_3DHRS,  d3_hrs),
            (COL_2D,      d2,       COL_2DHRS,  d2_hrs),
        ]

        for date_col, date_val, hrs_col, hrs_val in stage_data:
            state = classify_stage(date_val, hrs_val)
            st_bg, st_dark = STAGE_COLORS[state]
            st_fg = "1A1A1A" if st_dark else "FFFFFF"
            st_fill = _fill(st_bg)

            # Date range cell
            wc(date_col, date_val,
               font  = _font(size=9, color=st_fg),
               align = _align(h="center", v="center", wrap=True),
               fill  = st_fill)

            # Hours cell — show "—" for empty stages, bold for non-empty
            hrs_display = hrs_val if state != "EMPTY" else "—"
            hrs_bold    = state in ("DONE", "IN_PROGRESS") and hrs_val not in ("—", "")
            wc(hrs_col, hrs_display,
               font  = _font(bold=hrs_bold, size=9, color=st_fg),
               align = _align(h="center", v="center"),
               fill  = st_fill)

        # Total hrs
        wc(COL_TOTAL, total,
           align = _align(h="center", v="center"))

        # Designer
        wc(COL_DESIGNER, raw["designer"] or None,
           align = _align(h="left", v="center"))

        # Part Image (ONLY image_1_url)
        ref_url = raw["ref_url"]
        ref_cell = wc(COL_REFIMG, "View" if ref_url else None,
                    align=_align(h="center", v="center"))
        if ref_url:
            ref_cell.hyperlink = ref_url

        # Fix Image (ONLY image_2_url)
        fix_url = raw["work_url"]
        fix_cell = wc(COL_FIXIMG, "View" if fix_url else None,
                    align=_align(h="center", v="center"))
        if fix_url:
            fix_cell.hyperlink = fix_url

        # Work Proof — hyperlink
        proof_url  = raw["proof_url"]
        proof_cell = wc(COL_PROOF, "View" if proof_url else None,
                        align = _align(h="center", v="center"))
        if proof_url:
            proof_cell.hyperlink = proof_url

        ws.row_dimensions[r].height = 22

    # ── Summary / total row ────────────────────────────────────────────────
    total_row = 3 + display_row + 1
    STATUS_ORDER = ["Overdue", "In Progress", "On Hold", "Rework", "Review", "Assigned", "Closed"]
    count_parts  = [f"{status_counts[s]} {s}" for s in STATUS_ORDER if status_counts.get(s)]
    fixture_word = "fixture" if display_row == 1 else "fixtures"
    summary_text = f"TOTAL: {display_row} {fixture_word}  |  " + "  ·  ".join(count_parts)

    ws.merge_cells(f"A{total_row}:{last_col}{total_row}")
    tc            = ws.cell(total_row, 1, summary_text)
    tc.fill       = _fill(C_TOTAL_BG)
    tc.font       = _font(bold=True, size=9, color=C_TOTAL_FG)
    tc.alignment  = _align(h="left", v="center")
    tc.border     = Border(top=Side(style="medium", color=C_TOTAL_BG))
    ws.row_dimensions[total_row].height = 22

    # ── Freeze: first 3 columns + first 3 rows locked ─────────────────────
    ws.freeze_panes = "D4"

    # ── Legend sheet ──────────────────────────────────────────────────────
    _build_legend(wb)

    return wb


# ---------------------------------------------------------------------------
# Legend sheet
# ---------------------------------------------------------------------------

def _build_legend(wb: Workbook) -> None:
    ws = wb.create_sheet("Legend")

    ws.column_dimensions["A"].width = 3.0
    ws.column_dimensions["B"].width = 22.0
    ws.column_dimensions["C"].width = 35.0
    ws.column_dimensions["D"].width = 3.0
    ws.column_dimensions["E"].width = 20.0
    ws.column_dimensions["F"].width = 30.0

    # Title cells
    for col, title in [(2, "STATUS LEGEND"), (5, "STAGE PROGRESS LEGEND")]:
        c           = ws.cell(1, col, title)
        c.fill      = _fill(C_TITLE_BG)
        c.font      = _font(bold=True, size=12, color="FFFFFF")
        c.alignment = _align()

    # Status entries
    status_entries = [
        ("Assigned",    "Not yet started; queued for design"),
        ("In Progress", "Active work underway on one or more stages"),
        ("On Hold",     "Paused pending external dependency"),
        ("Review",      "All stages done; under QA / approval"),
        ("Rework",      "Returned for design correction"),
        ("Closed",      "Approved and completed"),
        ("Overdue",     "Deadline breached; escalation needed"),
    ]

    for i, (label, desc) in enumerate(status_entries, 2):
        bg, dark   = STATUS_COLORS[label]
        text_color = "1A1A1A" if dark else "FFFFFF"

        lc           = ws.cell(i, 2, label)
        lc.fill      = _fill(bg)
        lc.font      = _font(bold=True, size=9, color=text_color)
        lc.alignment = _align()

        dc           = ws.cell(i, 3, desc)
        dc.font      = _font(size=9, color="333333")
        dc.alignment = _align(h="left")

    # Stage progress entries
    stage_entries = [
        ("●  Completed",   "DONE",        "Stage done; duration logged"),
        ("◑  In Progress", "IN_PROGRESS", "Stage active; end date TBD"),
        ("○  Not Started", "EMPTY",       "Stage not yet triggered"),
    ]

    for i, (label, state_key, desc) in enumerate(stage_entries, 2):
        bg, dark   = STAGE_COLORS[state_key]
        text_color = "1A1A1A" if dark else "FFFFFF"

        lc           = ws.cell(i, 5, label)
        lc.fill      = _fill(bg)
        lc.font      = _font(bold=True, size=9, color=text_color)
        lc.alignment = _align()

        dc           = ws.cell(i, 6, desc)
        dc.font      = _font(size=9, color="333333")
        dc.alignment = _align(h="left")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    if len(sys.argv) != 3:
        print(
            f"Usage: python {sys.argv[0]} <input_raw.xlsx> <output_formatted.xlsx>",
            file=sys.stderr,
        )
        sys.exit(1)

    input_path  = sys.argv[1]
    output_path = sys.argv[2]

    wbs_title, raw_rows = read_raw(input_path)
    wb = build_formatted(wbs_title, raw_rows)
    wb.save(output_path)
    print(f"Formatted report written to: {output_path}")


if __name__ == "__main__":
    main()